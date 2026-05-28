#!/usr/bin/env python3
"""
Import individual student verification codes from the school Excel file.

Expected columns:
  연번, 현학번, 이름, 개별인증코드

The student id format is 5 digits: grade + 2-digit class + 2-digit number.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
import urllib.error
import urllib.request
from pathlib import Path

try:
    import openpyxl
except ImportError as exc:
    raise SystemExit("openpyxl is required to read .xlsx files.") from exc


REQUIRED_COLUMNS = ["현학번", "이름", "개별인증코드"]
DEFAULT_ENV = Path(__file__).resolve().parents[1] / ".env.local"
DEFAULT_SHEET = "인증코드"


def load_env(path: Path) -> dict[str, str]:
    env: dict[str, str] = {}
    if path.exists():
        for line in path.read_text(encoding="utf-8").splitlines():
            if not line or line.lstrip().startswith("#") or "=" not in line:
                continue
            key, value = line.split("=", 1)
            value = value.strip().strip('"').strip("'")
            env[key.strip()] = value
    return {**env, **os.environ}


def normalize_code(value: object) -> str:
    return re.sub(r"\s+", "", str(value or "")).upper()


def parse_student_id(value: object) -> tuple[str, int, int]:
    digits = re.sub(r"\D", "", str(value or ""))
    if len(digits) != 5:
        raise ValueError(f"현학번은 5자리 숫자여야 합니다: {value!r}")

    grade = int(digits[0])
    class_no = int(digits[1:3])
    if grade not in (1, 2, 3):
        raise ValueError(f"학년 범위가 올바르지 않습니다: {value!r}")
    if class_no < 1 or class_no > 11:
        raise ValueError(f"반 범위가 올바르지 않습니다: {value!r}")
    return digits, grade, class_no


def read_rows(path: Path, sheet_name: str) -> list[dict[str, object]]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    if sheet_name not in workbook.sheetnames:
        raise SystemExit(f"'{sheet_name}' 시트를 찾지 못했습니다. 시트: {', '.join(workbook.sheetnames)}")

    sheet = workbook[sheet_name]
    header = [str(cell.value or "").strip() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
    missing = [name for name in REQUIRED_COLUMNS if name not in header]
    if missing:
        raise SystemExit(f"필수 컬럼이 없습니다: {', '.join(missing)}")

    index = {name: header.index(name) for name in REQUIRED_COLUMNS}
    rows: list[dict[str, object]] = []
    seen_codes: set[str] = set()

    for row_no, cells in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if all(value is None or str(value).strip() == "" for value in cells):
            continue

        student_id = cells[index["현학번"]]
        student_name = str(cells[index["이름"]] or "").strip()
        code = normalize_code(cells[index["개별인증코드"]])
        student_no, grade, class_no = parse_student_id(student_id)

        if not student_name:
            raise ValueError(f"{row_no}행 이름이 비어 있습니다.")
        if len(code) < 4 or len(code) > 40:
            raise ValueError(f"{row_no}행 인증코드 길이가 올바르지 않습니다.")
        if code in seen_codes:
            raise ValueError(f"{row_no}행 인증코드가 중복입니다.")
        seen_codes.add(code)

        rows.append(
            {
                "code": code,
                "student_no": student_no,
                "student_name": student_name,
                "grade": grade,
                "class_no": class_no,
                "active": True,
            }
        )

    return rows


def supabase_upsert(url: str, service_key: str, rows: list[dict[str, object]], batch_size: int) -> None:
    endpoint = f"{url.rstrip('/')}/rest/v1/student_verification_codes?on_conflict=code"
    headers = {
        "apikey": service_key,
        "authorization": f"Bearer {service_key}",
        "content-type": "application/json",
        "prefer": "resolution=merge-duplicates,return=minimal",
    }

    for start in range(0, len(rows), batch_size):
        batch = rows[start : start + batch_size]
        req = urllib.request.Request(
            endpoint,
            data=json.dumps(batch, ensure_ascii=False).encode("utf-8"),
            headers=headers,
            method="POST",
        )
        try:
            with urllib.request.urlopen(req, timeout=30) as res:
                if res.status not in (200, 201, 204):
                    raise RuntimeError(f"Unexpected status: {res.status}")
        except urllib.error.HTTPError as exc:
            body = exc.read().decode("utf-8", errors="replace")
            raise RuntimeError(f"Supabase import failed ({exc.code}): {body}") from exc


def sql_literal(value: object) -> str:
    if value is True:
        return "true"
    if value is False:
        return "false"
    if isinstance(value, int):
        return str(value)
    return "'" + str(value).replace("'", "''") + "'"


def write_sql_import(path: Path, rows: list[dict[str, object]]) -> None:
    values = []
    for row in rows:
        values.append(
            "("
            + ", ".join(
                [
                    sql_literal(row["code"]),
                    sql_literal(row["student_no"]),
                    sql_literal(row["student_name"]),
                    sql_literal(row["grade"]),
                    sql_literal(row["class_no"]),
                    sql_literal(row["active"]),
                ]
            )
            + ")"
        )

    joined_values = ",\n".join(values)
    sql = f"""-- Individual student verification code import.
-- Generated from the local Excel file. Keep this file private.

begin;

create extension if not exists pgcrypto;

create table if not exists public.student_verification_codes (
  id uuid primary key default gen_random_uuid(),
  code text not null unique,
  student_no text,
  student_name text not null,
  grade integer not null check (grade between 1 and 3),
  class_no integer not null check (class_no between 1 and 11),
  active boolean not null default true,
  used_by uuid unique references auth.users(id) on delete set null,
  used_at timestamptz,
  created_at timestamptz not null default now(),
  updated_at timestamptz not null default now()
);

alter table public.student_verification_codes enable row level security;

create index if not exists student_verification_codes_code_idx
  on public.student_verification_codes (code);

create index if not exists student_verification_codes_used_by_idx
  on public.student_verification_codes (used_by);

alter table public.profiles
  add column if not exists student_verified boolean not null default false,
  add column if not exists student_no text,
  add column if not exists student_name text,
  add column if not exists student_verified_at timestamptz,
  add column if not exists verification_code_id uuid references public.student_verification_codes(id) on delete set null;

insert into public.student_verification_codes (code, student_no, student_name, grade, class_no, active)
values
{joined_values}
on conflict (code) do update set
  student_no = excluded.student_no,
  student_name = excluded.student_name,
  grade = excluded.grade,
  class_no = excluded.class_no,
  active = excluded.active,
  updated_at = now();

update public.profiles p
set
  student_no = v.student_no,
  student_name = coalesce(p.student_name, v.student_name),
  grade = coalesce(p.grade, v.grade),
  class_no = coalesce(p.class_no, v.class_no)
from public.student_verification_codes v
where v.used_by = p.id
  and v.student_no is not null;

commit;
"""
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text(sql, encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("xlsx", type=Path)
    parser.add_argument("--sheet", default=DEFAULT_SHEET)
    parser.add_argument("--env", type=Path, default=DEFAULT_ENV)
    parser.add_argument("--batch-size", type=int, default=200)
    parser.add_argument("--dry-run", action="store_true")
    parser.add_argument("--sql-output", type=Path)
    args = parser.parse_args()

    env = load_env(args.env)
    rows = read_rows(args.xlsx, args.sheet)

    grade_counts: dict[int, int] = {}
    for row in rows:
        grade = int(row["grade"])
        grade_counts[grade] = grade_counts.get(grade, 0) + 1

    print(f"validated_rows={len(rows)}")
    print("grade_counts=" + json.dumps(dict(sorted(grade_counts.items())), ensure_ascii=False))

    if args.dry_run:
        print("dry_run=true")
        return 0

    if args.sql_output:
        write_sql_import(args.sql_output, rows)
        print(f"sql_output={args.sql_output}")
        return 0

    url = env.get("NEXT_PUBLIC_SUPABASE_URL") or env.get("SUPABASE_URL")
    service_key = env.get("SUPABASE_SERVICE_ROLE_KEY")
    if not url or not service_key:
        raise SystemExit("Missing NEXT_PUBLIC_SUPABASE_URL/SUPABASE_URL or SUPABASE_SERVICE_ROLE_KEY")

    supabase_upsert(url, service_key, rows, args.batch_size)
    print(f"imported_rows={len(rows)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
